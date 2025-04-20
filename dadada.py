import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color, PatternFill, Border, Side, Alignment

data= {
    "Boca": {
        "ITCS103": 90,
        "ITPS102": 85,
    },
    "Oknarp": {
        "ITCS103": 95,
        "ITPS102": 80,
    },
    "Baldorado": {
        "ITCS103": 85,
        "ITPS102": 90,
    },

}

wb=Workbook()
sheet=wb.active
sheet.title="Student Grades"

sheet.title="Grades"

headings=["Name"] + list(data["Boca"].keys())
sheet.append(headings)

for student in data:
    grades=list(data[student].values())
    sheet.append([student] + list(grades))

for col in range(2, len(data["Boca"]) + 2):
    char=get_column_letter(col)
    sheet[char + "5"]= f"=SUM({char + '2'}:{char + '4'})/{len(data)}"

wb.save("grades.xlsx")



