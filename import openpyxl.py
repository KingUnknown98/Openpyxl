import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font 

data={
  "Boca": {
    "IT": 95,
    "ITNapay": 100,
    "ITmang": 0, 
  },
  "oknarp": {
    "IT": 95,
    "ITNapay": 101,
    "ITmang": 200,
  }
}

wb=Workbook()
ws=wb.active
ws.title="Grades ng IT na malupet"

headings=["Name"] + list(data["Boca"].keys())
ws.append(headings)

for student in data:
  Grades=list(data[student].values())
  ws.append([student] + list(Grades))
  
wb.save("dadac.xlsx")

